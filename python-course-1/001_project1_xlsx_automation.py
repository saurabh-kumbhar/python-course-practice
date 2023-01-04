import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference, Series


def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    corrected_price_title = sheet.cell(1,4)
    corrected_price_title.value = "corrected price"

    for row in range(2, sheet.max_row + 1):
        price_cell = sheet.cell(row, 3)
        corrected_price = price_cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    data = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    titles = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=sheet.max_row)

    chart = BarChart3D()
    chart.title = "Corrected Price per Transaction"
    chart.add_data(data)
    chart.set_categories(titles)

    sheet.add_chart(chart, "A7")

    wb.save(filename)


process_workbook("transactions.xlsx")